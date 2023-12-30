from django.shortcuts import render
from openpyxl import load_workbook

def get_books(genre=None):
    file_path = 'media/books.xlsx'
    workbook = load_workbook(filename=file_path)
    worksheet = workbook.active

    books = []
    skip_first_row = True
    for row in worksheet.iter_rows(values_only=True, min_row=2 if skip_first_row else 1):
        genre_check = row[3].lower() == genre.lower() if genre else True
        if genre_check:
            book = {
               'book_id': int(row[0]),           
               'title': row[1],
               'author': row[2],
               'genre': row[3],
               'year': int(row[4]),
               'description': row[5],
               'cover': row[6]
           }
            books.append(book)
    return books

def all_books(request):
    books = get_books()
    return render(request, 'books.html', {'books': books})

def book_detail(request, book_id):
    books = get_books()
    book = books[int(book_id) - 1] 
    return render(request, 'book_detail.html', {'book': book})

def thriller(request):
    books = get_books('триллер')
    return render(request, 'books.html', {'books': books})

def mystery(request):
    books = get_books('детектив')
    return render(request, 'books.html', {'books': books})

def fantasy(request):
    books = get_books('фэнтези')
    return render(request, 'books.html', {'books': books})

def programming(request):
    books = get_books('программирование')
    return render(request, 'books.html', {'books': books})