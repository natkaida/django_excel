from django.shortcuts import render
from openpyxl import load_workbook

def student_list(request):
    file_path = 'media/students.xlsx'
    workbook = load_workbook(filename=file_path)
    worksheet = workbook.active

    students = []
    skip_first_row = True
    for row in worksheet.iter_rows(values_only=True, min_row=2 if skip_first_row else 1):
        student = {
            'name': row[0],
            'lastname': row[1],
            'age': row[2],
            'faculty': row[3],
            'year': row[4],
            'average': row[5]
        }
        students.append(student)

    return render(request, 'student_list.html', {'students': students})
